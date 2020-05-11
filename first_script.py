from selenium import webdriver
from time import sleep
from Admin_creds import username,password    
from selenium.webdriver.common.by import By
#from usercreds import userid,uname,passd
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys  
import openpyxl
#import XLUtils
        

class ameyo():
    def __init__(self):
       self.driver = webdriver.Chrome()
        
    def login(self):
        

    
    #workbook=openpyxl.load_workbook(path)
    #sheet=workbook.active #workbook.get_sheet_by_name("sheet1")
        
    #rows = sheet.max_row #34
    #cols = sheet.max_column #4

    
    #print(rows)
    #print(cols)

    #for r in range (1,rows+1):

     #   for c in range (1,cols+1):
      #      print(sheet.cell(row=r,column=c).value,end="   ")
    
       # print()

        url = input("Enter Ameyo url: ")
        self.driver.get(url)
        sleep(2)

        email = self.driver.find_element_by_xpath('//*[@id="gwt-uid-1"]')
        email.send_keys(username)

        passd = self.driver.find_element_by_xpath('//*[@id="gwt-uid-2"]')
        passd.send_keys(password)

        clickbtn = self.driver.find_element_by_xpath('//*[@id="ameyo-body"]/div[3]/div/div[1]/div/div/div[2]/div/div[1]/div/div[2]/div/div[1]/div/div[2]/form/div[7]/div/button')
        clickbtn.click()

        sleep(10)
    def ad(self):

        user = self.driver.find_element_by_xpath('//*[@id="automationIdAdminUserManagementTab"]/span') 
        user.click()

        add = self.driver.find_element_by_xpath('//*[@id="automationIdAdminAddUserButton"]/span')
        add.click()

    def loop(self):

        path="/home/deadman/Downloads/Ameyo/automation/saved.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        rows=sheet.max_row
        cols=sheet.max_column
        


        for r in range (2,rows+1):
            #user_id=XLUtils.readData(path,r,1)
            user_id=sheet.cell(row=r,column=1).value
            user_name=sheet.cell(row=r,column=2).value
            pd=sheet.cell(row=r,column=3).value
            cpd=sheet.cell(row=r,column=4).value


        #self.driver.find_e

            uid = self.driver.find_element_by_xpath('//*[@id="gwt-uid-377"]')
            uid.send_keys(user_id)

        #actions = ActionChains(self)
        #actions.move_to_element(uid).click()
        #uid.send_keys("myme")

            usrname = self.driver.find_element_by_xpath('//*[@id="gwt-uid-379"]')
            usrname.send_keys(user_name)

            passwd = self.driver.find_element_by_xpath('//*[@id="gwt-uid-380"]')
            passwd.send_keys(pd)

            cpasswd = self.driver.find_element_by_xpath('//*[@id="gwt-uid-381"]')
            cpasswd.send_keys(cpd)

            sleep(7)



        #usertype = self.driver.find_element_by_xpath('//*[@id="select2-p26o-container"]')
        #usertype.click()

    #def  fp(self):
     #   pa = self.driver.find_element_by_xpath('//*[@id="gwt-uid-372"]/span[1]/span[1]/span/span[2]')
     #   drp=select(pa)
      #  drp.select_by_visibile_text('Professional-Agent')
            save = self.driver.find_element_by_xpath('//*[@id="automationIdAdminAddUserApplyButton"]')
            save.click()

            #if self.title=="Ameyo":
            #    print("it has passed")

           # print("it is working")
            #for d in range (2,5):


            user = self.driver.find_element_by_xpath('//*[@id="automationIdAdminUserManagementTab"]/span') 
            user.click()

            add = self.driver.find_element_by_xpath('//*[@id="automationIdAdminAddUserButton"]/span')
            add.click()

            sheet.cell(row=r,column=5).value="Success"
            
        workbook.save(path)
            #else:
             #   XLUtils.writeData(path,"Sheet1",r,5,"Failed")
        
            

    
    
    def logout(self):
        dd = self.driver.find_element_by_xpath('//*[@id="ameyo-body"]/div[4]/header/nav/div/ul[2]/li/a/i')
        dd.click()
        
        self.driver.find_element_by_xpath('//*[@id="userdp2"]/li[3]/a').click() 
        


 

bot = ameyo()
bot.login()
#bot.logout()
bot.ad()
bot.loop()  
