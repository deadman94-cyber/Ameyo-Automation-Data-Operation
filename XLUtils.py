import openpyxl

def getRowCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return(sheet.max_row)

def getColumnCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return(sheet.max_column)

def readData(file,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)




